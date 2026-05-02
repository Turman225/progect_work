import streamlit as st
import requests
import json
import re
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="🧪 Test Case Generator",
    page_icon="🧪",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
.metric-card {
    background: #f8f9fa;
    border-radius: 10px;
    padding: 16px 20px;
    text-align: center;
    border: 1px solid #e9ecef;
}
.metric-val { font-size: 32px; font-weight: 700; margin: 0; }
.metric-lbl { font-size: 13px; color: #6c757d; margin: 0; }
</style>
""", unsafe_allow_html=True)

# ── API URL ──────────────────────────────────────────────────────────────────
api_url = st.secrets.get("COLAB_API_URL", "") if hasattr(st, "secrets") else ""

# Всегда показываем поле ввода в сайдбаре
with st.sidebar:
    st.title("🧪 Test Case Generator")
    st.caption("Powered by Qwen2.5-VL-3B в Colab")
    st.divider()
    st.subheader("🔗 Подключение к Colab")
    api_url_input = st.text_input(
        "Colab API URL",
        value=api_url,
        placeholder="https://xxxx.trycloudflare.com",
        help="Скопируй URL из последней ячейки ноутбука в Colab",
    )
    if api_url_input:
        api_url = api_url_input.strip().rstrip("/")
    if api_url:
        try:
            r = requests.get(f"{api_url}/health", timeout=5)
            if r.status_code == 200:
                st.success("✅ Модель подключена")
                st.caption(r.json().get("model", ""))
            else:
                st.error("❌ Модель не отвечает")
        except Exception:
            st.error("❌ Нет соединения")
    else:
        st.warning("⬆️ Введи URL выше")
    st.divider()

    st.subheader("⚙️ Настройки")
    n_cases = st.slider("Количество тест-кейсов", 5, 20, 10)

    st.markdown("**Типы сценариев**")
    col_a, col_b, col_c = st.columns(3)
    with col_a: inc_pos = st.checkbox("✅ Pos",  value=True)
    with col_b: inc_neg = st.checkbox("❌ Neg",  value=True)
    with col_c: inc_bnd = st.checkbox("⚠️ Bnd", value=True)

    st.markdown("**Приоритеты**")
    col_d, col_e, col_f = st.columns(3)
    with col_d: inc_hi  = st.checkbox("🔴 High",   value=True)
    with col_e: inc_med = st.checkbox("🟡 Med",    value=True)
    with col_f: inc_lo  = st.checkbox("🟢 Low",    value=True)

    lang = st.selectbox("Язык", ["Russian", "English"])
    st.divider()
    st.caption("Model: Qwen2.5-VL-3B-Instruct")

if not api_url:
    st.title("🧪 Test Case Generator")
    st.info("👈 Вставь Colab API URL в боковой панели")
    st.markdown("""
    **Как получить URL:**
    1. Открой ноутбук `HW25_testcase_generator.ipynb` в Colab
    2. Запусти все ячейки по порядку
    3. В последней ячейке появится ссылка вида `https://xxxx.trycloudflare.com`
    4. Вставь её сюда
    """)
    st.stop()

# ── HELPERS ───────────────────────────────────────────────────────────────────
def to_str(v):
    if isinstance(v, list): return ", ".join(str(x) for x in v)
    return str(v) if v is not None else ""

def generate_test_cases(tz_text: str) -> list[dict]:
    payload = {"tz_text": tz_text, "n_cases": n_cases, "language": lang}
    try:
        r = requests.post(
            f"{api_url}/generate",
            json=payload,
            timeout=180,  # модель может генерировать 1-2 минуты
        )
        r.raise_for_status()
        return r.json().get("test_cases", [])
    except requests.exceptions.Timeout:
        st.error("Timeout — модель генерирует слишком долго. Уменьши количество тест-кейсов.")
        return []
    except requests.exceptions.ConnectionError:
        st.error("Нет соединения с Colab. Проверь что туннель запущен.")
        return []
    except Exception as e:
        st.error(f"Ошибка: {e}")
        return []

def to_excel(tcs: list[dict]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    headers = ["ID","Title","Type","Priority","Preconditions","Steps","Expected Result","Tags"]
    hf = PatternFill("solid", fgColor="1F4E79")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hf
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    colors = {"positive":"E8F5E9","negative":"FFEBEE","boundary":"FFF8E1"}
    for row, tc in enumerate(tcs, 2):
        fill = PatternFill("solid", fgColor=colors.get(tc.get("type",""),"FFFFFF"))
        steps_raw = tc.get("steps", [])
        if isinstance(steps_raw, str): steps_raw = [steps_raw]
        steps_text = "\n".join(f"{i+1}. {s}" for i, s in enumerate(steps_raw))
        vals = [to_str(tc.get("id")), to_str(tc.get("title")), to_str(tc.get("type")),
                to_str(tc.get("priority")), to_str(tc.get("preconditions")),
                steps_text, to_str(tc.get("expected_result")),
                ", ".join(str(t) for t in (tc.get("tags") or []))]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col, w in zip(range(1,9), [10,35,12,10,30,50,35,20]):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── MAIN ──────────────────────────────────────────────────────────────────────
st.title("🧪 Test Case Generator из ТЗ")
st.markdown("Загрузи ТЗ — Qwen2.5-VL-3B сгенерирует тест-кейсы")

tab1, tab2 = st.tabs(["✏️ Текст", "📁 Файл (PDF / DOCX)"])
tz_text = ""

with tab1:
    inp = st.text_area("Текст ТЗ:", height=280,
                        placeholder="Вставьте текст технического задания...")
    if inp:
        tz_text = inp
        st.caption(f"{len(tz_text):,} символов")

with tab2:
    uploaded = st.file_uploader("PDF или DOCX", type=["pdf","docx","doc"])
    if uploaded:
        import tempfile
        from pathlib import Path
        suf = Path(uploaded.name).suffix.lower()
        with tempfile.NamedTemporaryFile(delete=False, suffix=suf) as tmp:
            tmp.write(uploaded.read()); tmp_path = tmp.name
        try:
            if suf == ".pdf":
                import fitz
                doc = fitz.open(tmp_path)
                tz_text = "\n\n".join(p.get_text() for p in doc)
                doc.close()
            elif suf in (".docx", ".doc"):
                from docx import Document
                doc = Document(tmp_path)
                tz_text = "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
            st.success(f"✅ {uploaded.name} — {len(tz_text):,} символов")
            with st.expander("Просмотр"):
                st.text(tz_text[:2000] + ("..." if len(tz_text) > 2000 else ""))
        except Exception as e:
            st.error(f"Ошибка: {e}")

st.divider()
if st.button("🚀 Генерировать тест-кейсы", type="primary",
             use_container_width=True, disabled=not bool(tz_text)):
    with st.spinner("Qwen2.5-VL-3B генерирует тест-кейсы... (~1-2 мин)"):
        tcs = generate_test_cases(tz_text)
        if tcs:
            st.session_state["tcs"] = tcs
            st.success(f"✅ Сгенерировано {len(tcs)} тест-кейсов")
        else:
            st.error("Не удалось сгенерировать. Проверь соединение с Colab.")

if "tcs" in st.session_state and st.session_state["tcs"]:
    all_tcs = st.session_state["tcs"]
    active_types = [t for t,on in [("positive",inc_pos),("negative",inc_neg),("boundary",inc_bnd)] if on]
    active_prios = [p for p,on in [("high",inc_hi),("medium",inc_med),("low",inc_lo)] if on]
    tcs = [tc for tc in all_tcs if tc.get("type") in active_types and tc.get("priority") in active_prios]

    st.divider()
    c1,c2,c3,c4 = st.columns(4)
    with c1: st.markdown(f'<div class="metric-card"><p class="metric-val">{len(all_tcs)}</p><p class="metric-lbl">Всего</p></div>', unsafe_allow_html=True)
    with c2: st.markdown(f'<div class="metric-card"><p class="metric-val" style="color:#28a745">{sum(1 for t in all_tcs if t.get("type")=="positive")}</p><p class="metric-lbl">✅ Positive</p></div>', unsafe_allow_html=True)
    with c3: st.markdown(f'<div class="metric-card"><p class="metric-val" style="color:#dc3545">{sum(1 for t in all_tcs if t.get("type")=="negative")}</p><p class="metric-lbl">❌ Negative</p></div>', unsafe_allow_html=True)
    with c4: st.markdown(f'<div class="metric-card"><p class="metric-val" style="color:#fd7e14">{sum(1 for t in all_tcs if t.get("type")=="boundary")}</p><p class="metric-lbl">⚠️ Boundary</p></div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        st.download_button("📥 Excel (Zephyr/Jira)", to_excel(tcs), "test_cases.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
    with col2:
        st.download_button("📥 JSON (Zephyr API)", json.dumps(tcs, ensure_ascii=False, indent=2),
                           "test_cases.json", "application/json", use_container_width=True)

    st.subheader(f"📋 Тест-кейсы ({len(tcs)} из {len(all_tcs)})")
    ti = {"positive":"✅","negative":"❌","boundary":"⚠️"}
    pi = {"high":"🔴","medium":"🟡","low":"🟢"}
    for tc in tcs:
        label = f"{ti.get(tc.get('type',''),'?')} **{tc.get('id','?')}** — {tc.get('title','?')}  {pi.get(tc.get('priority','medium'),'⚪')}"
        with st.expander(label):
            l, r = st.columns([1,2])
            with l:
                st.markdown(f"**Тип:** `{tc.get('type')}`")
                st.markdown(f"**Приоритет:** `{tc.get('priority')}`")
                st.markdown(f"**Предусловия:**  \n{to_str(tc.get('preconditions','—'))}")
                if tc.get("tags"):
                    st.markdown("**Теги:** " + " ".join(f"`{t}`" for t in tc["tags"]))
            with r:
                st.markdown("**Шаги:**")
                steps = tc.get("steps",[])
                if isinstance(steps, str): steps = [steps]
                for idx, step in enumerate(steps, 1):
                    st.markdown(f"{idx}. {step}")
                st.success(f"**Ожидаемый результат:**  \n{to_str(tc.get('expected_result','—'))}")
elif tz_text:
    st.info("👆 Нажми «Генерировать тест-кейсы»")
else:
    st.info("👆 Введи текст ТЗ или загрузи файл")
